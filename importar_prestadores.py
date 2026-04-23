import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import load_workbook

CREDENCIAL   = "serviceAccountKey.json"
ARQUIVO_XLSX = "prestadores_rev1.xlsx"

cred = credentials.Certificate(CREDENCIAL)
firebase_admin.initialize_app(cred)
db = firestore.client()

wb = load_workbook(ARQUIVO_XLSX, data_only=True)
ws = wb["Prestadores"]

# Colunas: ID | Nome Atual | Nome Corrigido | Quem indicou | Observação | Categoria | Status | Telefone | WhatsApp
#           0        1             2                3              4           5          6        7          8

atualizados = 0
ignorados   = 0
erros       = 0

print("Processando planilha...\n")

for row in ws.iter_rows(min_row=2, values_only=True):
    doc_id         = str(row[0] or "").strip()
    nome_corrigido = str(row[2] or "").strip()
    indicacao      = str(row[3] or "").strip()
    observacao     = str(row[4] or "").strip()
    categoria      = str(row[5] or "").strip()

    if not doc_id:
        continue

    if not nome_corrigido:
        print(f"  [IGNORADO] ID {doc_id} — 'Nome Corrigido' está vazio.")
        ignorados += 1
        continue

    try:
        db.collection("providers").document(doc_id).update({
            "name":        nome_corrigido,
            "categoryName": categoria,
            "indicacao":   indicacao,
            "observacao":  observacao,
        })
        print(f"  [OK] {nome_corrigido} | cat: {categoria or '—'} | ind: {indicacao or '—'} | obs: {observacao or '—'}")
        atualizados += 1
    except Exception as e:
        print(f"  [ERRO] ID {doc_id}: {e}")
        erros += 1

print(f"\nConcluído: {atualizados} atualizados | {ignorados} ignorados | {erros} erros.")

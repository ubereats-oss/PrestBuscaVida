import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CREDENCIAL    = "serviceAccountKey.json"
ARQUIVO_SAIDA = "prestadores.xlsx"

cred = credentials.Certificate(CREDENCIAL)
firebase_admin.initialize_app(cred)
db = firestore.client()

print("Buscando categorias...")
categorias = {}
for doc in db.collection("categories").stream():
    categorias[doc.id] = doc.to_dict().get("name", "")
print(f"{len(categorias)} categorias carregadas.")

print("Buscando prestadores...")
registros = []
for doc in db.collection("providers").stream():
    d = doc.to_dict()
    category_id   = d.get("categoryId", "")
    category_name = d.get("categoryName", "") or categorias.get(category_id, "")
    registros.append({
        "id":             doc.id,
        "name":           d.get("name", ""),
        "name_corrigido": d.get("name", ""),
        "categoryName":   category_name,
        "status":         d.get("status", ""),
        "phone":          d.get("phone", ""),
        "whatsapp":       d.get("whatsapp", ""),
    })

registros.sort(key=lambda x: (x["categoryName"].lower(), x["name"].lower()))
print(f"{len(registros)} prestadores encontrados.")

wb = Workbook()
ws = wb.active
ws.title = "Prestadores"

cabecalhos = ["ID (não editar)", "Nome Atual", "Nome Corrigido ✏️", "Categoria", "Status", "Telefone", "WhatsApp"]
larguras   = [30, 35, 35, 22, 12, 18, 18]

header_fill   = PatternFill("solid", start_color="1F3864")
editable_fill = PatternFill("solid", start_color="FFF2CC")
locked_fill   = PatternFill("solid", start_color="F2F2F2")
border_side   = Side(style="thin", color="CCCCCC")
border        = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

for col, (cab, larg) in enumerate(zip(cabecalhos, larguras), 1):
    cell = ws.cell(row=1, column=col, value=cab)
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border
    ws.column_dimensions[get_column_letter(col)].width = larg

ws.row_dimensions[1].height = 30

for row_idx, r in enumerate(registros, 2):
    valores = [r["id"], r["name"], r["name_corrigido"], r["categoryName"], r["status"], r["phone"], r["whatsapp"]]
    for col, val in enumerate(valores, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center")
        cell.border    = border
        if col == 3:
            cell.fill = editable_fill
            cell.font = Font(name="Arial", size=10, bold=True)
        else:
            cell.fill = locked_fill

ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:G{len(registros) + 1}"

instrucoes = wb.create_sheet("Instruções")
instrucoes["A1"] = "INSTRUÇÕES DE USO"
instrucoes["A1"].font = Font(bold=True, size=14, name="Arial")
instrucoes["A2"] = ""
instrucoes["A3"] = "1. Edite APENAS a coluna 'Nome Corrigido ✏️' (fundo amarelo) na aba 'Prestadores'."
instrucoes["A4"] = "2. Não altere nem apague a coluna 'ID (não editar)' — ela é usada para atualizar o registro correto."
instrucoes["A5"] = "3. Salve o arquivo como .xlsx."
instrucoes["A6"] = "4. Execute o script importar_prestadores.py para aplicar as correções."
for row in range(1, 7):
    instrucoes.cell(row=row, column=1).font = Font(name="Arial", size=11)
instrucoes.column_dimensions["A"].width = 80

wb.save(ARQUIVO_SAIDA)
print(f"\nArquivo '{ARQUIVO_SAIDA}' gerado com sucesso.")

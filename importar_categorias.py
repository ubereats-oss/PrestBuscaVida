import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import load_workbook

CREDENCIAL   = "serviceAccountKey.json"
ARQUIVO_XLSX = "categorias.xlsx"

cred = credentials.Certificate(CREDENCIAL)
firebase_admin.initialize_app(cred)
db = firestore.client()

wb = load_workbook(ARQUIVO_XLSX, data_only=True)
ws = wb["Categorias"]

# Colunas: ID | Nome Atual | Nome Corrigido | Slug | Ordem | Ativo
#           0       1             2            3      4       5

atualizados = 0
ignorados   = 0
erros       = 0

print("Processando planilha...\n")

for row in ws.iter_rows(min_row=2, values_only=True):
    doc_id         = str(row[0] or "").strip()
    nome_atual     = str(row[1] or "").strip()
    nome_corrigido = str(row[2] or "").strip()

    if not doc_id:
        continue

    if not nome_corrigido:
        print(f"  [IGNORADO] ID {doc_id} — 'Nome Corrigido' está vazio.")
        ignorados += 1
        continue

    if nome_atual == nome_corrigido:
        ignorados += 1
        continue

    try:
        db.collection("categories").document(doc_id).update({"name": nome_corrigido})
        print(f"  [OK] '{nome_atual}'  →  '{nome_corrigido}'")
        atualizados += 1
    except Exception as e:
        print(f"  [ERRO] ID {doc_id}: {e}")
        erros += 1

print(f"\nConcluído: {atualizados} atualizadas | {ignorados} sem alteração | {erros} erros.")
